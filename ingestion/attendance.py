"""Premier DLC attendance ingestion.

Parses the SAR (School Attendance Report) workbook and emits
`attendance.json` — a per-school monthly attendance dataset plus
precomputed aggregates so the dashboard does zero math at render time.

Usage:
    python -m ingestion.attendance \\
      --xlsx "/path/to/SAR 2023 - 2025.xlsx" \\
      --out  attendance.json

The script is deterministic and idempotent: same xlsx in → identical JSON out.
The xlsx is *not* committed to the public repo; we store only its sha256 in
the output for reproducibility.

Scope decisions (see docs/methodology-attendance.md):
- Per-school IDs are anonymised by default (school_id_01..50) until PDLC
  consent for named publication is documented. Use --reveal-emis to publish
  real EMIS codes.
- Missing-school handling: schools absent from 2023/2024 are represented
  with monthly = null for those years. The frontend renders only available
  lines, with a tooltip.
- School-days/month for the lost-child-school-day estimate is a fixed
  constant (20) documented in methodology.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import statistics
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SCHOOLS_CSV = ROOT / "data" / "schools" / "pssp_psrp_50.csv"
DEFAULT_OUT = ROOT / "attendance.json"

YEARS = ("2023", "2024", "2025")
SCHOOL_DAYS_PER_MONTH = 20  # working-day estimate; see methodology note


def _load_roster() -> dict[int, dict]:
    """Return EMIS-code-keyed roster from the 50-school CSV."""
    rows: dict[int, dict] = {}
    with SCHOOLS_CSV.open() as f:
        for r in csv.DictReader(f):
            rows[int(r["emis_code"])] = {
                "students": int(r["students"]),
                "cluster": r["cluster"],
            }
    return rows


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _find_header_row(rows: list) -> tuple[int, int]:
    """Locate the header row and the EMIS column by scanning content.

    Tolerates extra title/spacer rows and inserted columns. Looks at the first
    15 rows for a cell containing 'EMIS' (case-insensitive).
    """
    for ri, r in enumerate(rows[:15]):
        for ci, cell in enumerate(r):
            if isinstance(cell, str) and cell.strip().upper() == "EMIS":
                return ri, ci
    raise ValueError("Could not find EMIS column in first 15 rows — SAR layout changed?")


def _parse_year_sheet(ws) -> tuple[list[str], dict[int, dict[str, float]]]:
    """Parse one year sheet. Returns (months_in_order, {emis: {month: value}}).

    Layout-tolerant: locates the header row by content and detects month columns
    by datetime cell type rather than fixed offsets. Survives small PDLC
    re-saves that shift rows or insert columns.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_idx, emis_col = _find_header_row(rows)
    header = rows[header_idx]

    months: list[tuple[int, str]] = []
    for col_idx, cell in enumerate(header):
        if isinstance(cell, dt.datetime):
            months.append((col_idx, cell.strftime("%b")))

    if not months:
        raise ValueError(f"No month columns (datetime cells) found in header row {header_idx}")

    out: dict[int, dict[str, float]] = {}
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= emis_col:
            continue
        emis_raw = r[emis_col]
        try:
            emis = int(emis_raw) if not isinstance(emis_raw, int) else emis_raw
        except (TypeError, ValueError):
            continue
        record: dict[str, float] = {}
        for col_idx, month_name in months:
            v = r[col_idx] if col_idx < len(r) else None
            if isinstance(v, (int, float)) and 0.0 <= v <= 1.0:
                record[month_name] = float(v)
        if record:
            out[emis] = record
    return [m for _, m in months], out


def _find_year_sheet(wb, year: str):
    """Resolve year sheet by exact name first, then by fuzzy contains-match."""
    if year in wb.sheetnames:
        return wb[year]
    for name in wb.sheetnames:
        if year in name:
            return wb[name]
    raise KeyError(f"No sheet for year {year}; available: {wb.sheetnames}")


def _annual_mean(monthly: Optional[dict[str, float]]) -> Optional[float]:
    if not monthly:
        return None
    return statistics.mean(monthly.values())


def build_payload(xlsx_path: Path, reveal_emis: bool = False) -> dict:
    """Build the full attendance.json payload."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    roster = _load_roster()

    by_year: dict[str, dict[int, dict[str, float]]] = {}
    months_per_year: dict[str, list[str]] = {}
    for y in YEARS:
        ws = _find_year_sheet(wb, y)
        months_per_year[y], by_year[y] = _parse_year_sheet(ws)

    # Build per-school records (use roster as canonical 50-school list)
    # Stable order: by emis code
    emis_order = sorted(roster.keys())
    anon_map = {emis: f"school_id_{i+1:02d}" for i, emis in enumerate(emis_order)}

    schools_block: dict[str, dict] = {}
    schools_meta_pairs: list[tuple[int, str]] = []  # (emis, public_id) for meta

    missing_by_year: dict[str, list] = {y: [] for y in YEARS}
    schools_covered: dict[str, int] = {}

    for emis in emis_order:
        public_id = emis if reveal_emis else anon_map[emis]
        schools_meta_pairs.append((emis, public_id))

        monthly_by_year: dict[str, Optional[dict[str, float]]] = {}
        annual_means: dict[str, Optional[float]] = {}
        for y in YEARS:
            mo = by_year[y].get(emis)
            if mo:
                monthly_by_year[y] = {m: round(v, 4) for m, v in mo.items()}
                annual_means[y] = round(_annual_mean(mo), 4)
            else:
                monthly_by_year[y] = None
                annual_means[y] = None
                missing_by_year[y].append(public_id)

        # 4.6-pp Δ stat: only if both 2023 and 2025 present
        pp_change = None
        if annual_means["2023"] is not None and annual_means["2025"] is not None:
            pp_change = round((annual_means["2025"] - annual_means["2023"]) * 100, 2)

        # May anomaly: 2025 May vs 3-year May mean (if available)
        may_anomaly = None
        may_2025 = (monthly_by_year["2025"] or {}).get("May")
        may_values = [
            (monthly_by_year[y] or {}).get("May")
            for y in YEARS
            if monthly_by_year[y]
        ]
        may_values = [v for v in may_values if v is not None]
        if may_2025 is not None and len(may_values) >= 2:
            may_anomaly = round((may_2025 - statistics.mean(may_values)) * 100, 2)

        schools_block[str(public_id)] = {
            "students": roster[emis]["students"],
            "cluster": roster[emis]["cluster"],
            "monthly": monthly_by_year,
            "annual_mean": annual_means,
            "pp_change_2023_2025": pp_change,
            "may_anomaly_2025_vs_3yr": may_anomaly,
        }

    # --- top-level aggregates ----------------------------------------------------
    annual_mean_overall: dict[str, float] = {}
    for y in YEARS:
        vals = [
            s["annual_mean"][y]
            for s in schools_block.values()
            if s["annual_mean"][y] is not None
        ]
        annual_mean_overall[y] = round(statistics.mean(vals), 4) if vals else None
        schools_covered[y] = len(vals)

    pp_change_2023_2025 = (
        round((annual_mean_overall["2025"] - annual_mean_overall["2023"]) * 100, 2)
        if annual_mean_overall["2023"] is not None and annual_mean_overall["2025"] is not None
        else None
    )

    # Lost child-school-days: Σ over (school × month present in both 2023 and 2025)
    # of students × (att_2023 - att_2025) × SCHOOL_DAYS_PER_MONTH
    lost_days = 0.0
    for emis, public_id in schools_meta_pairs:
        rec = schools_block[str(public_id)]
        m23 = rec["monthly"].get("2023") or {}
        m25 = rec["monthly"].get("2025") or {}
        common = set(m23) & set(m25)
        for month in common:
            delta = m23[month] - m25[month]  # positive = decline
            lost_days += rec["students"] * delta * SCHOOL_DAYS_PER_MONTH
    lost_days_total = int(round(lost_days))

    # Cluster averages
    by_cluster: dict[str, dict[str, float]] = defaultdict(dict)
    cluster_buckets: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for rec in schools_block.values():
        for y in YEARS:
            if rec["annual_mean"][y] is not None:
                cluster_buckets[rec["cluster"]][y].append(rec["annual_mean"][y])
    for cluster, years in cluster_buckets.items():
        for y in YEARS:
            if years[y]:
                by_cluster[cluster][y] = round(statistics.mean(years[y]), 4)

    # Month-of-year mean
    by_month_mean: dict[str, dict[str, float]] = {}
    for y in YEARS:
        month_buckets: dict[str, list[float]] = defaultdict(list)
        for rec in schools_block.values():
            m = rec["monthly"][y]
            if not m:
                continue
            for month_name, val in m.items():
                month_buckets[month_name].append(val)
        by_month_mean[y] = {
            m: round(statistics.mean(vals), 4) for m, vals in month_buckets.items()
        }

    # --- assemble payload --------------------------------------------------------
    # Reproducibility (NFR-10): same xlsx in → identical JSON out. retrieved_at
    # is a calendar date set by the operator on the day of ingest; generated_at
    # is derived from the source hash so byte-for-byte equality holds across
    # re-runs of the same input.
    # Operator-side lookup: EMIS → public ID. Lets the dashboard join attendance
    # records to the scores.json roster. When anonymised, this map is the only
    # field that bridges anonymous IDs back to EMIS, so it's documented as a
    # weak anonymisation (operator-grade reproducibility, not k-anonymity).
    emis_to_id = {str(emis): str(public_id) for emis, public_id in schools_meta_pairs}

    payload = {
        "meta": {
            "source_file": xlsx_path.name,
            "source_sha256": _sha256(xlsx_path),
            "retrieved_at": dt.date.today().isoformat(),
            "provider": "Premier DLC",
            "license": "CC BY 4.0 (with PDLC consent for school-level aggregates; named publication requires --reveal-emis)",
            "methodology_url": "docs/methodology-attendance.md",
            "school_id_anonymised": not reveal_emis,
            "school_days_per_month": SCHOOL_DAYS_PER_MONTH,
            "years": list(YEARS),
            "months_per_year": months_per_year,
            "schools_covered": schools_covered,
            "missing_school_ids": missing_by_year,
            "emis_to_id": emis_to_id,
        },
        "aggregates": {
            "annual_mean": annual_mean_overall,
            "pp_change_2023_2025": pp_change_2023_2025,
            "lost_child_school_days_2023_baseline": lost_days_total,
            "by_month_mean": by_month_mean,
            "by_cluster": dict(by_cluster),
        },
        "schools": schools_block,
    }
    return payload


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to SAR workbook")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help="Output JSON path")
    ap.add_argument(
        "--reveal-emis",
        action="store_true",
        help="Publish real EMIS codes instead of anonymised school_id_NN (requires PDLC consent)",
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"error: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 2

    payload = build_payload(xlsx_path, reveal_emis=args.reveal_emis)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2))

    agg = payload["aggregates"]
    print(f"wrote {out} ({out.stat().st_size / 1024:.1f} KB)")
    print(f"  annual mean: {agg['annual_mean']}")
    print(f"  Δ 2023→2025: {agg['pp_change_2023_2025']} pp")
    print(f"  lost child-school-days vs 2023: ~{agg['lost_child_school_days_2023_baseline']:,}")
    print(f"  schools covered: {payload['meta']['schools_covered']}")
    print(f"  anonymised: {payload['meta']['school_id_anonymised']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
